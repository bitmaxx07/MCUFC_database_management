import random

import jwt
import openpyxl
import qrcode

mDict = {
    'a': 37,
    'b': 39,
    'c': 65,
    'd': 54,
    'e': 32,
    'f': 76,
    'g': 61,
    'h': 87,
    'i': 83,
    'j': 11,
    'k': 12,
    'l': 57,
    'm': 41,
    'n': 91,
    'o': 38,
    'p': 69,
    'q': 29,
    'r': 13,
    's': 74,
    't': 45,
    'u': 55,
    'v': 28,
    'w': 94,
    'x': 97,
    'y': 25,
    'z': 20,
    'A': 37,
    'B': 39,
    'C': 65,
    'D': 54,
    'E': 32,
    'F': 76,
    'G': 61,
    'H': 87,
    'I': 83,
    'J': 11,
    'K': 12,
    'L': 57,
    'M': 41,
    'N': 91,
    'O': 38,
    'P': 69,
    'Q': 29,
    'R': 13,
    'S': 74,
    'T': 45,
    'U': 55,
    'V': 28,
    'W': 94,
    'X': 97,
    'Y': 25,
    'Z': 20,
    '0': 8,
    '1': 4,
    '2': 6,
    '3': 7,
    '4': 3,
    '5': 5,
    '6': 0,
    '7': 2,
    '8': 9,
    '9': 1}

# TODO: CONFIGURE HERE!!!
filename = "会员卡_20230820_程琦_2.xlsx"

key = "secret"
wb = openpyxl.load_workbook(filename)

col_lastname = 1
col_firstname = 2
col_birthday = 3
col_id = 4
col_subname = 6
col_website = 5
col_num = 7


def generate_random_number():
    random_number = random.randint(12, 99)
    if random_number % 10 != 1:
        return random_number


def assign_id(worksheet):
    num_set = set()
    start_row = 0
    for row in range(1, worksheet.max_row + 1):
        # print(worksheet.cell(row, col_num).value)
        if worksheet.cell(row, col_num).value != "" and worksheet.cell(row, col_num).value is not None:
            num_set.add(worksheet.cell(row, col_num).value)
        if worksheet.cell(row, col_num).value == "" or worksheet.cell(row, col_num).value is None:
            start_row = row
            break

    def check_duplicate_num(check_str):
        res = check_str
        if check_str in num_set:
            res = check_str[:-2] + str(generate_random_number())
            check_duplicate_num(res)
        # if check_str not in num_set:
        return res

    if worksheet.title == "Fussball Mitgliederlist":
        for row in range(start_row, worksheet.max_row + 1):
            checkstring = "01" + worksheet.cell(row, col_lastname).value[0] + worksheet.cell(row, col_firstname).value[0] \
                          + worksheet.cell(row, col_birthday).value.strftime('%y%m%d')
            '''if checkstring + "00" in num_set:
                while checkstring + str(generate_random_number()) in num_set:
                    worksheet.cell(row, col_num).value = checkstring + str(generate_random_number())
            else:
                worksheet.cell(row, col_num).value = checkstring + "00"'''
            checkstring = check_duplicate_num(checkstring + "00")
            print(checkstring)
            num_set.add(checkstring)
            worksheet.cell(row, col_num).value = checkstring

            print(worksheet.cell(row, col_lastname).value)
            print(worksheet.cell(row, col_firstname).value)
            print(worksheet.cell(row, col_num).value)
            code_string = worksheet.cell(row, col_lastname).value + "_" + \
                          worksheet.cell(row, col_firstname).value + "_" + worksheet.cell(row, col_num).value
            encoded = jwt.encode({code_string: ""}, key, algorithm="HS256")
            worksheet.cell(row, col_id).value = encoded
            worksheet.cell(row, col_subname).value = encoded.replace(".", "-")[-75:].lower()
            worksheet.cell(row, col_website).value = "https://www.csm-ev.com/members/football/" + encoded.replace(".", "-")[
                                                                                           -75:].lower()
            worksheet.cell(row, 8).value = ''.join(str(mDict[c]) for c in ws.cell(row, col_num).value)

            input_data = "https://www.csm-ev.com/members/football/" + encoded.replace(".", "-")[-75:].lower()
            qr = qrcode.QRCode(version=1, box_size=10, border=0)
            qr.add_data(input_data)
            qr.make(fit=True)
            img = qr.make_image(fill="black", back_color="white")
            img.save("qrcodes/" + code_string + ".png")
            print("saved QR code for " + code_string)
            print("------------------------")

    if worksheet.title == "VIP":
        start_row = 0
        for row in range(1, worksheet.max_row + 1):
            if worksheet.cell(row, col_num).value != "" and worksheet.cell(row, col_num).value is not None:
                num_set.add(worksheet.cell(row, col_num).value)

            if worksheet.cell(row, col_num).value == "" or worksheet.cell(row, col_num).value is None:
                start_row = row
                break

        for row in range(start_row, worksheet.max_row + 1):
            checkstring = "00" + worksheet.cell(row, col_lastname).value[0] + \
                          worksheet.cell(row, col_firstname).value[0] + "000000"
            '''if checkstring + "00" in num_set:
                while checkstring + str(generate_random_number()) not in num_set:
                    worksheet.cell(row, col_num).value = checkstring + str(generate_random_number())
            else:
                worksheet.cell(row, col_num).value = checkstring + "00"'''
            checkstring = check_duplicate_num(checkstring + "00")
            print(checkstring)
            num_set.add(checkstring)
            worksheet.cell(row, col_num).value = checkstring

            print(worksheet.cell(row, col_lastname).value)
            print(worksheet.cell(row, col_firstname).value)
            print(worksheet.cell(row, col_num).value)

            code_string = worksheet.cell(row, col_lastname).value + "_" + worksheet.cell(row, col_firstname).value \
                          + "_" + worksheet.cell(row, col_num).value
            encoded = jwt.encode({code_string: ""}, key, algorithm="HS256")
            worksheet.cell(row, col_id).value = encoded
            worksheet.cell(row, col_subname).value = encoded.replace(".", "-")[-75:].lower()
            worksheet.cell(row, col_website).value = "https://www.csm-ev.com/members/VIP/" + \
                                                     encoded.replace(".", "-")[-75:].lower()
            worksheet.cell(row, 8).value = ''.join(str(mDict[c]) for c in ws.cell(row, col_num).value)

            input_data = "https://www.csm-ev.com/members/VIP/" + encoded.replace(".", "-")[-75:].lower()
            qr = qrcode.QRCode(version=1, box_size=10, border=0)
            qr.add_data(input_data)
            qr.make(fit=True)
            img = qr.make_image(fill="black", back_color="white")
            img.save("qrcodes_vip/" + worksheet.cell(row, col_lastname).value + "_" +
                     worksheet.cell(row, col_firstname).value + "_" + worksheet.cell(row, 8).value + ".png")
            print("saved QR code for " + code_string)
            print("------------------------")


temp = input("f for football, v for vip: ")

if temp == "f":
    ws = wb["Fussball Mitgliederlist"]
    assign_id(ws)

elif temp == "v":
    ws = wb["VIP"]
    assign_id(ws)

wb.save(filename)
