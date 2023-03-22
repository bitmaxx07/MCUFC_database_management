import jwt
import openpyxl
import qrcode
import PIL
from hashlib import sha256

mDict = {
    'a': 37,
    'b': 39,
    'c': 65,
    'd': 54,
    'e': 32,
    'f': 76,
    'g': 61,
    'h': 87,
    'i': 83,
    'j': 11,
    'k': 12,
    'l': 57,
    'm': 41,
    'n': 91,
    'o': 38,
    'p': 69,
    'q': 29,
    'r': 13,
    's': 74,
    't': 45,
    'u': 55,
    'v': 28,
    'w': 94,
    'x': 97,
    'y': 25,
    'z': 20,
    'A': 37,
    'B': 39,
    'C': 65,
    'D': 54,
    'E': 32,
    'F': 76,
    'G': 61,
    'H': 87,
    'I': 83,
    'J': 11,
    'K': 12,
    'L': 57,
    'M': 41,
    'N': 91,
    'O': 38,
    'P': 69,
    'Q': 29,
    'R': 13,
    'S': 74,
    'T': 45,
    'U': 55,
    'V': 28,
    'W': 94,
    'X': 97,
    'Y': 25,
    'Z': 20,
    '0': 8,
    '1': 4,
    '2': 6,
    '3': 7,
    '4': 3,
    '5': 5,
    '6': 0,
    '7': 2,
    '8': 9,
    '9': 1}

key = "secret"
# encoded = jwt.encode({"chengqi": ""}, key, algorithm="HS256")
# print(encoded.replace(".", "-"))
# print(encoded.replace(".", "-")[0: 75].lower())

# print(jwt.decode(encoded, key, algorithms="HS256"))

id_dict = {}

wb = openpyxl.load_workbook("会员卡_20230314.xlsx")
# CONFIGURE HERE
ws = wb["Fussball Mitgliederlist"]
wb_id = openpyxl.load_workbook("MemberID.xlsx")
ws_id = wb_id.active

col_lastname = 1
col_firstname = 2
col_birthday = 3
col_id = 4
col_subname = 6
col_website = 5
col_num = 7

col_checknum = 5

'''
ID rule:
id = col_lastname.value[0] + col_firstname.value[0] + birthday.strftime('%y%m%d')
'''

id_mapping = {}
for row in range(1, ws.max_row + 1):
    for row_id in range(1, ws_id.max_row + 1):
        print(ws_id.cell(row_id, col_id).value)
        print(ws_id.cell(row_id, col_checknum).value)
        id_mapping.update({ws_id.cell(row_id, col_id).value: ws_id.cell(row_id, col_checknum).value})
    # fill in excel sheet and generate id mapping
    # ws.cell(row, col_num).value = "0" + str(1000000 + row)
    print(id_mapping)
    # CONFIGURE HERE
    checkstring = "01" + ws.cell(row, col_lastname).value[0] + ws.cell(row, col_firstname).value[0] \
                  + ws.cell(row, col_birthday).value.strftime('%y%m%d')

    for k in id_mapping.keys():
        if checkstring == k:
            for r in range(2, ws_id.max_row + 1):
                if ws_id.cell(r, col_id).value == checkstring:
                    if ws_id.cell(r, col_lastname).value == ws.cell(row, col_lastname).value and ws_id.cell(r,
                                                                                                            col_firstname).value == ws.cell(
                            row, col_firstname).value:
                        pass
                    else:
                        ws_id.cell(r, col_checknum).value += 1

    # POTENTIAL PROBLEM HERE!!!
    if checkstring in id_mapping.keys():
        ws.cell(row, col_num).value = checkstring + "{:02d}".format(id_mapping[checkstring])
    else:
        ws.cell(row, col_num).value = checkstring + "00"

    code_string = ws.cell(row, col_lastname).value + "_" + \
                  ws.cell(row, col_firstname).value + "_" + ws.cell(row, col_num).value
    encoded = jwt.encode({code_string: ""}, key, algorithm="HS256")
    ws.cell(row, col_id).value = encoded
    ws.cell(row, col_subname).value = encoded.replace(".", "-")[-75:].lower()
    id_dict.update({code_string: encoded.replace(".", "-")[-75:].lower()})
    ws.cell(row, col_website).value = "https://www.csm-ev.com/members/football/" + encoded.replace(".", "-")[
                                                                                   -75:].lower()
    ws.cell(row, 8).value = ''.join(str(mDict[c]) for c in ws.cell(row, col_num).value)

    name_list = []
    for r in range(2, ws_id.max_row + 1):
        name_list.append(ws_id.cell(r, 1).value + ws_id.cell(r, 2).value)
    if ws.cell(row, col_lastname).value + ws.cell(row, col_firstname).value not in name_list:
        temp_maxrow = ws_id.max_row + 1
        ws_id.cell(temp_maxrow, col_lastname).value = ws.cell(row, col_lastname).value
        ws_id.cell(temp_maxrow, col_firstname).value = ws.cell(row, col_firstname).value
        ws_id.cell(temp_maxrow, col_birthday).value = str(ws.cell(row, col_birthday).value.strftime('%y%m%d'))
        ws_id.cell(temp_maxrow, col_id).value = checkstring
        wb_id.save("MemberID.xlsx")

    print("name/personal ID: " + code_string)
    print("id: " + ws.cell(row, col_id).value)
    print("subname: " + ws.cell(row, col_subname).value)
    print("website: " + ws.cell(row, col_website).value)
    print("----------------------------")

    # generate qr codes
    input_data = "https://www.csm-ev.com/members/football/" + encoded.replace(".", "-")[-75:].lower()
    qr = qrcode.QRCode(version=1, box_size=10, border=0)
    qr.add_data(input_data)
    qr.make(fit=True)

    img = qr.make_image(fill="black", back_color="white")
    img.save("qrcodes/" + code_string + ".png")
    print("saved QR code for " + code_string)
    print("------------------------")

wb.save("会员卡_20220314.xlsx")
print("-----------FINISHED----------")
print("TOTAL MAPPING:")
for key in id_dict.keys():
    print("name: " + key + " id: " + id_dict.get(key))
