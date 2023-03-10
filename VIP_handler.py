import jwt
import openpyxl
import qrcode
import PIL

key = "VIP"

id_dict = {}

wb = openpyxl.load_workbook("会员卡.xlsx")
ws = wb["VIP"]

col_lastname = 1
col_firstname = 2
col_id = 3
col_subname = 5
col_website = 4
col_num = 6

for row in range(1, ws.max_row + 1):
    # fill in excel sheet and generate id mapping
    ws.cell(row, col_num).value = "0" + str(2000000 + row)
    code_string = ws.cell(row, col_lastname).value + "_" + \
                  ws.cell(row, col_firstname).value + "_" + \
                  str(ws.cell(row, col_num).value)
    encoded = jwt.encode({code_string: ""}, key, algorithm="HS256")
    ws.cell(row, col_id).value = encoded
    ws.cell(row, col_subname).value = encoded.replace(".", "-")[0: 75].lower()
    id_dict.update({code_string: encoded.replace(".", "-")[0: 75].lower()})
    ws.cell(row, col_website).value = "https://www.csm-ev.com/members/vip/" + encoded.replace(".", "-")[0: 75].lower()
    print("name/personal ID: " + code_string)
    print("id: " + ws.cell(row, col_id).value)
    print("subname: " + ws.cell(row, col_subname).value)
    print("website: " + ws.cell(row, col_website).value)
    print("personal ID: " + ws.cell(row, col_num).value)
    print("----------------------------")

    # generate qr codes
    input_data = "https://www.csm-ev.com/members/vip/" + encoded.replace(".", "-")[0: 75].lower()
    qr = qrcode.QRCode(version=1, box_size=10, border=0)
    qr.add_data(input_data)
    qr.make(fit=True)

    img = qr.make_image(fill="black", back_color="white")
    img.save("qrcodes/" + code_string + ".png")
    print("saved QR code for " + code_string)
    print("------------------------")


wb.save("会员卡.xlsx")
print("-----------FINISHED----------")
print("TOTAL MAPPING:")
for key in id_dict.keys():
    print("name: " + key + " id: " + id_dict.get(key))
